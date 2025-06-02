import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk  # type: ignore
import unicodedata
import os
from openpyxl import Workbook, load_workbook  # type: ignore
from datetime import datetime
import shutil

# Lista válida de prendas
PRENDAS_VALIDAS = [
    "camiseta", "pantalón", "falda", "abrigo", "chaqueta", "jersey",
    "sudadera", "camisa", "blusa", "vaqueros", "chaleco", "bikini",
    "traje", "corbata", "bufanda", "guantes", "calcetines", "ropa interior",
    "pijama", "shorts", "chanclas", "zapatos", "zapatillas", "botas"
]

def normalizar(texto):
    texto = texto.lower().strip()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    return texto

class GestorPrendasApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestor de Prendas - Tèxtils Montblanc")
        self.root.geometry("600x700")
        self.root.configure(bg="#eaf0f6")

        self.inventario = {}

        # Logo
        logo_path = "logo_montblanc.png"
        if os.path.exists(logo_path):
            logo_img = Image.open(logo_path)
            max_width, max_height = 300, 100
            ratio = min(max_width / logo_img.width, max_height / logo_img.height)
            new_size = (int(logo_img.width * ratio), int(logo_img.height * ratio))
            logo_img = logo_img.resize(new_size, Image.LANCZOS)
            self.logo = ImageTk.PhotoImage(logo_img)
            logo_label = tk.Label(root, image=self.logo, bg="#eaf0f6")
            logo_label.pack(pady=(15, 5))
        else:
            tk.Label(root, text="[Logo no encontrado]", bg="#eaf0f6", fg="red").pack(pady=(20, 10))

        # Título
        tk.Label(root, text="Gestor de Inventario", font=("Segoe UI", 20, "bold"), bg="#eaf0f6", fg="#1f3b73").pack(pady=(0, 20))

        # Marco de contenido
        marco = tk.Frame(root, bg="#ffffff", bd=2, relief="groove")
        marco.pack(padx=20, pady=10, fill="both", expand=False)

        # Selector de cantidad
        cantidad_frame = tk.Frame(marco, bg="#ffffff")
        cantidad_frame.pack(pady=(15, 5))
        tk.Label(cantidad_frame, text="Cantidad:", font=("Segoe UI", 12), bg="#ffffff").pack(side="left", padx=(0, 10))
        self.cantidad = tk.IntVar(value=1)
        self.spin = tk.Spinbox(cantidad_frame, from_=1, to=100, textvariable=self.cantidad, font=("Segoe UI", 12), width=5)
        self.spin.pack(side="left")

        # Entrada
        self.entrada = tk.Entry(marco, font=("Segoe UI", 12), width=30)
        self.entrada.pack(pady=10)
        self.entrada.bind("<Return>", self.anadir_prenda)

        # Botones
        boton_frame = tk.Frame(marco, bg="#ffffff")
        boton_frame.pack(pady=(10, 10))
        self.crear_boton(boton_frame, "➕ Añadir", self.anadir_prenda).pack(side="left", padx=5)
        self.crear_boton(boton_frame, "🗑️ Eliminar", self.eliminar_prenda).pack(side="left", padx=5)
        self.crear_boton(boton_frame, "💾 Guardar", self.guardar_en_excel).pack(side="left", padx=5)

        # Área de texto con scrollbar
        texto_frame = tk.Frame(marco, bg="#ffffff")
        texto_frame.pack(pady=15)
        self.texto = tk.Text(texto_frame, height=15, width=50, font=("Segoe UI", 10), bd=1, relief="solid")
        self.texto.pack(side="left")
        scrollbar = tk.Scrollbar(texto_frame, command=self.texto.yview)
        scrollbar.pack(side="right", fill="y")
        self.texto.config(yscrollcommand=scrollbar.set)

        self.actualizar_texto()

    def crear_boton(self, parent, texto, comando):
        return tk.Button(parent, text=texto, command=comando, font=("Segoe UI", 11, "bold"),
                         bg="#1f3b73", fg="white", activebackground="#294a8d",
                         activeforeground="white", relief="flat", padx=10, pady=5)

    def anadir_prenda(self, event=None):
        prenda_raw = self.entrada.get()
        cantidad = self.cantidad.get()
        prenda = normalizar(prenda_raw)

        if prenda in map(normalizar, PRENDAS_VALIDAS):
            prenda_formateada = next(p for p in PRENDAS_VALIDAS if normalizar(p) == prenda)
            self.inventario[prenda_formateada] = self.inventario.get(prenda_formateada, 0) + cantidad
            self.entrada.delete(0, tk.END)
            self.actualizar_texto()
        else:
            messagebox.showwarning("Prenda no válida", f"'{prenda_raw}' no es una prenda reconocida. Inténtalo de nuevo.")

    def eliminar_prenda(self):
        prenda_raw = self.entrada.get()
        prenda = normalizar(prenda_raw)

        for p in PRENDAS_VALIDAS:
            if normalizar(p) == prenda:
                if p in self.inventario:
                    del self.inventario[p]
                    self.entrada.delete(0, tk.END)
                    self.actualizar_texto()
                    return
                else:
                    messagebox.showinfo("No encontrada", f"'{p}' no está en el inventario.")
                    return

        messagebox.showwarning("Prenda no válida", f"'{prenda_raw}' no es una prenda reconocida.")

    def actualizar_texto(self):
        self.texto.delete("1.0", tk.END)
        if not self.inventario:
            self.texto.insert(tk.END, "No hay prendas añadidas.")
        else:
            for prenda, cantidad in self.inventario.items():
                self.texto.insert(tk.END, f"{prenda.title()}: {cantidad}\n")

    def guardar_en_excel(self):
        original = "inventario.xlsx"
        temporal = "inventario_temp.xlsx"
        hoja_nombre = "Historial"
        datos_totales = {}

        try:
            if os.path.exists(original):
                backup_name = f"inventario_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                shutil.copyfile(original, backup_name)

                wb = load_workbook(original)
                if hoja_nombre in wb.sheetnames:
                    ws = wb[hoja_nombre]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        fecha, prenda, cantidad = row
                        prenda_key = prenda.lower()
                        datos_totales[prenda_key] = [cantidad, fecha]
                else:
                    ws = wb.create_sheet(hoja_nombre)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = hoja_nombre
                ws.append(["Fecha y hora", "Prenda", "Cantidad"])

            fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for prenda, cantidad in self.inventario.items():
                prenda_key = prenda.lower()
                if prenda_key in datos_totales:
                    datos_totales[prenda_key][0] += cantidad
                    datos_totales[prenda_key][1] = fecha_hora
                else:
                    datos_totales[prenda_key] = [cantidad, fecha_hora]

            if hoja_nombre in wb.sheetnames:
                del wb[hoja_nombre]
            ws = wb.create_sheet(hoja_nombre)
            ws.append(["Fecha y hora", "Prenda", "Cantidad"])

            for prenda_key, (cantidad, fecha) in datos_totales.items():
                ws.append([fecha, prenda_key.title(), cantidad])

            wb.save(temporal)
            shutil.move(temporal, original)

            messagebox.showinfo("Guardado", "El inventario se ha guardado correctamente en 'inventario.xlsx'.")

        except PermissionError:
            if os.path.exists(temporal):
                os.remove(temporal)
            messagebox.showwarning(
                "Archivo en uso",
                "El archivo 'inventario.xlsx' está abierto en Excel.\nCierra el archivo y vuelve a intentarlo."
            )

# Ejecutar app
if __name__ == "__main__":
    root = tk.Tk()
    app = GestorPrendasApp(root)

    def on_close():
        if app.inventario:
            app.guardar_en_excel()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()
